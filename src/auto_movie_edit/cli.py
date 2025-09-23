"""Command line interface for the auto_movie_edit toolkit."""

from __future__ import annotations

import json
from hashlib import md5
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import typer
from openpyxl import Workbook, load_workbook

from .srt import SrtParseError, parse_srt
from .workbook import (
    DEFAULT_TEMPLATE,
    create_workbook_template,
    load_workbook_data,
    save_workbook,
)
from .ymmp import apply_hiragana_shrink, build_project, write_outputs

app = typer.Typer(help="Auto Movie Edit CLI utilities")


@app.command("make-sheet")
def make_sheet(
    srt: Path = typer.Option(..., exists=True, dir_okay=False, readable=True, help="SRT subtitle file"),
    out: Path = typer.Option(..., dir_okay=False, help="Output Excel file"),
) -> None:
    """Create a workbook template populated with SRT subtitles."""
    try:
        entries = parse_srt(srt)
    except SrtParseError as exc:
        typer.secho(str(exc), fg=typer.colors.RED)
        raise typer.Exit(code=1) from exc

    workbook = create_workbook_template(DEFAULT_TEMPLATE)
    timeline_sheet = workbook["TIMELINE"]

    for row_index, entry in enumerate(entries, start=2):
        timeline_sheet.cell(row=row_index, column=1, value=entry.start.to_string())
        timeline_sheet.cell(row=row_index, column=2, value=entry.end.to_string())
        timeline_sheet.cell(row=row_index, column=3, value=entry.text)

    save_workbook(workbook, out)
    typer.secho(f"Workbook created: {out}", fg=typer.colors.GREEN)


@app.command("build")
def build(
    sheet: Path = typer.Option(..., exists=True, dir_okay=False, readable=True, help="Timeline workbook"),
    out: Path = typer.Option(Path("work"), file_okay=False, dir_okay=True, help="Output directory"),
) -> None:
    """Build a simplified YMMP project from the workbook."""
    data = load_workbook_data(sheet)
    project, warnings = build_project(data)
    write_outputs(project, warnings, out)
    typer.secho(f"Project generated with {len(warnings)} warnings -> {out}", fg=typer.colors.GREEN)


@app.command("filter")
def filter_command(
    filter_name: str = typer.Argument(..., help="Filter name"),
    input_path: Path = typer.Option(..., exists=True, dir_okay=False, help="Input YMMP JSON"),
    out: Path = typer.Option(..., dir_okay=False, help="Output YMMP JSON"),
    scale: Optional[float] = typer.Option(0.85, help="Scale applied to telop text"),
) -> None:
    """Apply post-processing filters to a project."""
    filter_name = filter_name.lower()
    if filter_name == "hira-shrink":
        apply_hiragana_shrink(input_path, out, scale or 0.85)
        typer.secho(f"Hiragana shrink applied -> {out}", fg=typer.colors.GREEN)
    else:
        typer.secho(f"Unknown filter: {filter_name}", fg=typer.colors.RED)
        raise typer.Exit(code=1)


def _extract_telops_from_raw_ymmp(project: dict, xlsx_path: Path) -> dict:
    """Extracts TextItems and saves them as templates."""
    typer.secho("Extracting telop patterns...", fg=typer.colors.CYAN)
    extracted = {}
    template_dir = xlsx_path.parent / "templates" / "telops"
    template_dir.mkdir(parents=True, exist_ok=True)

    items = project.get("Timelines", [{}])[0].get("Items", [])
    for item in items:
        if "TextItem" in item.get("$type", ""):
            text = item.get("Text", "no_text")
            pattern_id = f"telop_{text}"
            if pattern_id in extracted: continue

            template_file_path = template_dir / f"{pattern_id}.json"
            template_file_path.write_text(json.dumps(item, ensure_ascii=False, indent=2), encoding="utf-8")

            extracted[pattern_id] = {
                "pattern_id": pattern_id,
                "source": template_file_path.resolve().as_posix(),
                "description": f"「{text}」から自動抽出",
            }
    typer.secho(f"Found {len(extracted)} telop patterns.", fg=typer.colors.GREEN)
    return extracted

def _extract_assets_from_raw_ymmp(project: dict, xlsx_path: Path) -> dict:
    """Extracts TachieItems/ImageItems and saves them as templates."""
    typer.secho("Extracting asset patterns...", fg=typer.colors.CYAN)
    extracted = {}
    template_dir = xlsx_path.parent / "templates" / "assets"
    template_dir.mkdir(parents=True, exist_ok=True)

    items = project.get("Timelines", [{}])[0].get("Items", [])
    for item in items:
        item_type = item.get("$type", "")
        asset_id = None
        kind = None

        if "TachieItem" in item_type:
            kind = "tachie"
            char_name = item.get("CharacterName", "unknown")
            # 表情をファイル名から推測
            eye_path = Path(item.get("TachieItemParameter", {}).get("Eye", ""))
            emotion = eye_path.stem.split("】")[-1] if "】" in eye_path.stem else "default"
            asset_id = f"tachie_{char_name}_{emotion}"
            
        elif "ImageItem" in item_type:
            kind = "image"
            file_path = Path(item.get("FilePath", ""))
            asset_id = f"image_{file_path.stem}"

        if asset_id and asset_id not in extracted:
            template_file_path = template_dir / f"{asset_id}.json"
            template_file_path.write_text(json.dumps(item, ensure_ascii=False, indent=2), encoding="utf-8")
            
            extracted[asset_id] = {
                "asset_id": asset_id,
                "kind": kind,
                "path": template_file_path.resolve().as_posix(),
                "notes": f"from {Path(project.get('FilePath')).name}",
            }
    typer.secho(f"Found {len(extracted)} asset patterns.", fg=typer.colors.GREEN)
    return extracted


def _extract_packs_from_raw_ymmp(project: dict, xlsx_path: Path) -> dict:
    """Extracts multi-object packs from timeline items."""
    typer.secho("Extracting pack patterns...", fg=typer.colors.CYAN)
    extracted: Dict[str, Dict[str, Any]] = {}
    template_dir = xlsx_path.parent / "templates" / "packs"
    template_dir.mkdir(parents=True, exist_ok=True)

    seen_hashes: set[str] = set()
    items = project.get("Timelines", [{}])[0].get("Items", [])
    for item in items:
        if not isinstance(item, dict):
            continue
        if not _looks_like_pack(item):
            continue

        sanitized = _strip_runtime_fields(item)
        digest = _hash_template(sanitized)
        if digest in seen_hashes:
            continue
        seen_hashes.add(digest)

        pack_id = f"pack_{digest[:8]}"
        template_path = template_dir / f"{_safe_filename(pack_id)}.json"
        template_path.write_text(json.dumps(sanitized, ensure_ascii=False, indent=2), encoding="utf-8")

        extracted[pack_id] = {
            "pack_id": pack_id,
            "source": _relative_template_path(template_path, xlsx_path.parent),
            "notes": f"Extracted from {Path(project.get('FilePath', 'source.ymmp')).name}",
        }

    typer.secho(f"Found {len(extracted)} pack patterns.", fg=typer.colors.GREEN)
    return extracted


def _extract_fx_from_raw_ymmp(project: dict, xlsx_path: Path, packs: Dict[str, Dict[str, Any]]) -> dict:
    """Extracts FX presets by mirroring effect items as packs."""
    typer.secho("Extracting FX presets...", fg=typer.colors.CYAN)
    extracted: Dict[str, Dict[str, Any]] = {}
    fx_dir = xlsx_path.parent / "templates" / "fx"
    fx_dir.mkdir(parents=True, exist_ok=True)
    pack_dir = xlsx_path.parent / "templates" / "packs"
    pack_dir.mkdir(parents=True, exist_ok=True)

    items = project.get("Timelines", [{}])[0].get("Items", [])
    for item in items:
        if not isinstance(item, dict):
            continue
        item_type = item.get("$type", "")
        if not _looks_like_fx(item_type):
            continue

        sanitized = _strip_runtime_fields(item)
        digest = _hash_template(sanitized)
        fx_id = f"fx_{digest[:8]}"
        if fx_id in extracted:
            continue

        fx_template_path = fx_dir / f"{_safe_filename(fx_id)}.json"
        fx_template_path.write_text(json.dumps(sanitized, ensure_ascii=False, indent=2), encoding="utf-8")

        pack_id = f"pack_{digest[:8]}"
        if pack_id not in packs:
            pack_template_path = pack_dir / f"{_safe_filename(pack_id)}.json"
            if not pack_template_path.exists():
                pack_template_path.write_text(json.dumps(sanitized, ensure_ascii=False, indent=2), encoding="utf-8")
            packs[pack_id] = {
                "pack_id": pack_id,
                "source": _relative_template_path(pack_template_path, xlsx_path.parent),
                "notes": "Auto-generated from FX item",
            }

        extracted[fx_id] = {
            "fx_id": fx_id,
            "fx_type": _infer_fx_type(item_type),
            "source": pack_id,
        }

    typer.secho(f"Found {len(extracted)} FX presets.", fg=typer.colors.GREEN)
    return extracted


@app.command("absorb")
def absorb(
    ymmp: Path = typer.Option(..., exists=True, dir_okay=False, readable=True, help="Source YMMP JSON"),
    xlsx: Path = typer.Option(..., dir_okay=False, help="Workbook to update"),
) -> None:
    """Absorb a YMMP file into the workbook dictionaries."""
    xlsx = xlsx.resolve()
    try:
        project = json.loads(ymmp.read_text(encoding="utf-8-sig"))
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        typer.secho(f"Error reading YMMP file: {e}", fg=typer.colors.RED)
        raise typer.Exit(code=1)

    if xlsx.exists():
        workbook = load_workbook(xlsx)
    else:
        workbook = create_workbook_template(DEFAULT_TEMPLATE)
        
    base_dir = xlsx.parent
    if (
        "telop_patterns" in project
        or "assets" in project
        or "packs" in project
        or "fx_presets" in project
    ):  # Tool-generated file
        telop_raw = project.get("telop_patterns", {})
        asset_raw = project.get("assets", {})
        pack_raw = project.get("packs", {})
        fx_raw = project.get("fx_presets", project.get("fx", {}))

        telop_patterns = telop_raw if isinstance(telop_raw, dict) else {}
        assets = asset_raw if isinstance(asset_raw, dict) else {}
        packs = pack_raw if isinstance(pack_raw, dict) else {}
        fx_presets = fx_raw if isinstance(fx_raw, dict) else {}
    else:  # Raw YMM4 file
        telop_patterns = _extract_telops_from_raw_ymmp(project, xlsx)
        assets = _extract_assets_from_raw_ymmp(project, xlsx)
        packs = _extract_packs_from_raw_ymmp(project, xlsx)
        fx_presets = _extract_fx_from_raw_ymmp(project, xlsx, packs)

    _prepare_templates_for_sync(base_dir, telop_patterns, assets, packs)

    _sync_dictionary_sheet(workbook, "TELP_PATTERNS", DEFAULT_TEMPLATE.telp_headers, telop_patterns)
    _sync_dictionary_sheet(workbook, "ASSETS_SINGLE", DEFAULT_TEMPLATE.asset_headers, assets)
    _sync_dictionary_sheet(workbook, "PACKS_MULTI", DEFAULT_TEMPLATE.pack_headers, packs)
    _sync_dictionary_sheet(workbook, "FX", DEFAULT_TEMPLATE.fx_headers, fx_presets)
    
    save_workbook(workbook, xlsx)
    typer.secho(f"Workbook updated with project dictionaries -> {xlsx}", fg=typer.colors.GREEN)


def _sync_dictionary_sheet(workbook: Workbook, name: str, headers: Iterable[str], values: dict) -> None:
    header_list = list(headers)
    sheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)

    if sheet.max_row < 1: _write_headers(sheet, header_list)
        
    existing_ids = {str(sheet.cell(row=r, column=1).value) for r in range(2, sheet.max_row + 1)}

    for key, payload in values.items():
        if str(key) in existing_ids: continue
        
        new_row_index = sheet.max_row + 1
        for col_idx, header in enumerate(header_list, start=1):
            field_name = _map_header_to_field(header)
            # Use key as the default for the first column's field name
            is_id_column = (field_name == header_list[0].lower().replace(" ", "_")) or \
                           (field_name in ["pattern_id", "asset_id", "pack_id", "fx_id"])
            
            cell_value = key if is_id_column else payload.get(field_name)

            if isinstance(cell_value, (dict, list)) and cell_value:
                cell_value = json.dumps(cell_value, ensure_ascii=False)
            elif isinstance(cell_value, (dict, list)):
                cell_value = None

            sheet.cell(row=new_row_index, column=col_idx, value=cell_value)
        existing_ids.add(str(key))


def _write_headers(sheet, headers: Iterable[str]) -> None:
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=header)


def _map_header_to_field(header: str) -> str:
    mapping = {
        "パターンID": "pattern_id", "参照ソース": "source", "上書きキー": "overrides",
        "説明": "description", "備考": "notes", "素材ID": "asset_id", "種別": "kind",
        "パス": "path", "既定レイヤ": "default_layer", "既定X": "default_x",
        "既定Y": "default_y", "既定ズーム": "default_zoom", "パックID": "pack_id",
        "役割": "role", "レイヤ帯": "layer", "FX_ID": "fx_id", "種類": "fx_type",
        "パック": "source", "アセット": "asset", "パラメータ": "parameters",
    }
    return mapping.get(header, header.lower().replace(" ", "_"))


def _prepare_templates_for_sync(
    base_dir: Path,
    telops: Dict[str, Dict[str, Any]],
    assets: Dict[str, Dict[str, Any]],
    packs: Dict[str, Dict[str, Any]],
) -> None:
    for key, payload in telops.items():
        _persist_template_payload(base_dir, "telops", key, payload, "overrides", "source")

    for key, payload in assets.items():
        _persist_template_payload(base_dir, "assets", key, payload, "parameters", "path")

    for key, payload in packs.items():
        _persist_template_payload(base_dir, "packs", key, payload, "overrides", "source")


def _persist_template_payload(
    base_dir: Path,
    category: str,
    identifier: str,
    payload: Dict[str, Any],
    data_field: str,
    path_field: str,
) -> None:
    if not isinstance(payload, dict):
        return

    template_data = payload.get(data_field)
    if not isinstance(template_data, (dict, list)) or not template_data:
        return

    template_dir = base_dir / "templates" / category
    template_dir.mkdir(parents=True, exist_ok=True)
    template_path = template_dir / f"{_safe_filename(identifier)}.json"
    template_path.write_text(json.dumps(template_data, ensure_ascii=False, indent=2), encoding="utf-8")

    relative_path = _relative_template_path(template_path, base_dir)
    payload[path_field] = relative_path
    payload[data_field] = None


def _relative_template_path(path: Path, base_dir: Path) -> str:
    try:
        return path.relative_to(base_dir).as_posix()
    except ValueError:
        return path.resolve().as_posix()


def _safe_filename(identifier: str) -> str:
    sanitized = identifier.replace("\\", "_").replace("/", "_").replace(":", "_")
    return sanitized or "template"


def _strip_runtime_fields(data: Any) -> Any:
    if isinstance(data, dict):
        return {
            key: _strip_runtime_fields(value)
            for key, value in data.items()
            if key not in {"Frame", "Length"}
        }
    if isinstance(data, list):
        return [_strip_runtime_fields(item) for item in data]
    return data


def _hash_template(data: Any) -> str:
    serialized = json.dumps(data, ensure_ascii=False, sort_keys=True)
    return md5(serialized.encode("utf-8")).hexdigest()


def _looks_like_pack(item: Dict[str, Any]) -> bool:
    if not isinstance(item, dict):
        return False
    if isinstance(item.get("Items"), list) and item["Items"]:
        return True
    item_type = item.get("$type", "")
    return any(keyword in item_type for keyword in ("Group", "Repeat", "Tachie"))


def _looks_like_fx(item_type: str) -> bool:
    if not isinstance(item_type, str):
        return False
    candidates = ("Effect", "Zoom", "Shake", "Blur", "Speedline", "Vignette", "Filter")
    return any(keyword.lower() in item_type.lower() for keyword in candidates)


def _infer_fx_type(item_type: str) -> str:
    if not item_type:
        return "fx"
    base = item_type.split(",")[0].split(".")[-1]
    return base.replace("Item", "").lower() or "fx"

if __name__ == "__main__":
    app()
