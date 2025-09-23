"""Functions for reading and writing workbook files used by the pipeline."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook, load_workbook

from .models import (
    Asset,
    Character,
    FxPreset,
    LayerBand,
    Pack,
    TelopPattern,
    TimelineFx,
    TimelineObject,
    TimelineRow,
    WorkbookData,
)
from .utils import ensure_list, iter_nonempty, parse_mapping, parse_timecode


TELP_HEADERS = [
    "パターンID", "参照ソース", "上書きキー", "基準幅", "基準高さ",
    "FPS", "説明", "備考",
]
ASSET_HEADERS = [
    "素材ID", "種別", "パス", "既定レイヤ", "既定X",
    "既定Y", "既定ズーム", "備考",
]
PACK_HEADERS = [
    "パックID", "参照ソース", "上書きキー", "基準幅", "基準高さ",
    "FPS", "備考",
]
LAYER_HEADERS = ["役割", "レイヤ帯"]
FX_HEADERS = ["FX_ID", "種類", "パック", "アセット", "パラメータ"]
TIMELINE_HEADERS = [
    "開始", "終了", "字幕テキスト", "テロップ", "キャラクター", "表情(3つ内包)",
    "表情", "他1", "パック", "オブジェクト1", "オブジェクト2", "オブジェクト3",
    "背景", "FX_PARAM", "承認", "メモ",
]
SCHEMA_HEADERS = ["シート", "日本語", "キー"]
CHARACTERS_HEADERS = ["キャラクター名", "パーツ名", "ベースパス"]


@dataclass(slots=True)
class WorkbookTemplate:
    telp_headers: List[str]; asset_headers: List[str]; pack_headers: List[str]
    layer_headers: List[str]; fx_headers: List[str]; timeline_headers: List[str]
    schema_headers: List[str]; characters_headers: List[str]

DEFAULT_TEMPLATE = WorkbookTemplate(
    telp_headers=TELP_HEADERS, asset_headers=ASSET_HEADERS, pack_headers=PACK_HEADERS,
    layer_headers=LAYER_HEADERS, fx_headers=FX_HEADERS, timeline_headers=TIMELINE_HEADERS,
    schema_headers=SCHEMA_HEADERS, characters_headers=CHARACTERS_HEADERS,
)


def _write_headers(sheet, headers: Iterable[str]) -> None:
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

def create_workbook_template(template: WorkbookTemplate = DEFAULT_TEMPLATE) -> Workbook:
    wb = Workbook(); wb.remove(wb.active)
    _write_headers(wb.create_sheet("TELP_PATTERNS"), template.telp_headers)
    _write_headers(wb.create_sheet("ASSETS_SINGLE"), template.asset_headers)
    _write_headers(wb.create_sheet("PACKS_MULTI"), template.pack_headers)
    _write_headers(wb.create_sheet("LAYERS"), template.layer_headers)
    _write_headers(wb.create_sheet("FX"), template.fx_headers)
    _write_headers(wb.create_sheet("TIMELINE"), template.timeline_headers)
    _write_headers(wb.create_sheet("SCHEMA_MAP"), template.schema_headers)
    _write_headers(wb.create_sheet("CHARACTERS"), template.characters_headers)
    return wb

def save_workbook(workbook: Workbook, path: str | Path) -> None:
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)

def load_sheet_dictionaries(sheet) -> list[dict[str, Any]]:
    if not sheet or sheet.max_row < 1: return []
    headers = [(cell.value or "").strip() for cell in sheet[1]]
    return [
        {h: row[i] for i, h in enumerate(headers) if h}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

def _split_schema_columns(raw_value: Any, fallback: str | None) -> List[str]:
    columns: List[str] = []
    if isinstance(raw_value, str):
        normalized = raw_value.replace(";", ",").replace("|", ",")
        columns = [chunk.strip() for chunk in normalized.split(",") if chunk.strip()]
    elif raw_value not in (None, ""):
        columns = [str(raw_value).strip()]
    if not columns and fallback:
        columns = [fallback]
    return columns

def _schema_columns(schema: Dict[str, List[str]], key: str, fallback: str | None) -> List[str]:
    columns = schema.get(key)
    if columns:
        return columns
    return [fallback] if fallback else []

def _single_column(schema: Dict[str, List[str]], key: str, fallback: str | None) -> str | None:
    columns = _schema_columns(schema, key, fallback)
    return columns[0] if columns else None

def _collect_schema_columns(schema: Dict[str, List[str]], prefix: str) -> List[str]:
    collected: List[tuple[str, List[str]]] = [
        (key, cols)
        for key, cols in schema.items()
        if key.startswith(prefix) and cols
    ]
    collected.sort(key=lambda item: item[0])
    result: List[str] = []
    seen: set[str] = set()
    for _, columns in collected:
        for column in columns:
            if column in seen:
                continue
            seen.add(column)
            result.append(column)
    return result

def _schema_column_key_map(schema: Dict[str, List[str]]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for key, columns in schema.items():
        for column in columns:
            mapping[column] = key
    return mapping

def _row_value(row: Dict[str, Any], schema: Dict[str, List[str]], key: str, fallback: str | None) -> Any:
    for column in _schema_columns(schema, key, fallback):
        if column in row:
            return row.get(column)
    return None

def _first_nonempty(row: Dict[str, Any], columns: List[str]) -> Any:
    for column in columns:
        value = row.get(column)
        if value not in (None, ""):
            return value
    return None

def _resolve_role_name(column: str, canonical_key: str | None) -> str:
    if not canonical_key:
        return column
    if canonical_key.startswith("object."):
        suffix = canonical_key.split(".", 1)[1]
        return suffix or column
    return canonical_key

def _resolve_layer_band(layers: Dict[str, LayerBand], column: str, canonical_key: str | None, role_name: str) -> int | None:
    candidates = []
    if canonical_key:
        if canonical_key.startswith("object."):
            candidates.append(role_name)
        else:
            candidates.append(canonical_key)
    candidates.append(role_name)
    candidates.append(column)
    for key in candidates:
        layer = layers.get(key)
        if layer:
            return layer.layer
    return None

def load_workbook_data(path: str | Path) -> WorkbookData:
    path = Path(path).resolve()
    wb = load_workbook(path, data_only=True)
    data = WorkbookData()

    # Schema map must be loaded first so that column lookups can be resolved.
    if "SCHEMA_MAP" in wb.sheetnames:
        for row in iter_nonempty(load_sheet_dictionaries(wb["SCHEMA_MAP"])):
            sheet_name = _string_or_none(row.get("シート"))
            key = _string_or_none(row.get("キー"))
            columns = _split_schema_columns(row.get("日本語"), key)
            if not sheet_name or not key or not columns:
                continue
            sheet_map = data.schema_map.setdefault(sheet_name, {})
            sheet_map[key] = columns

    # Load all dictionaries using the resolved schema map
    telp_schema = data.schema_map.get("TELP_PATTERNS", {})
    asset_schema = data.schema_map.get("ASSETS_SINGLE", {})
    pack_schema = data.schema_map.get("PACKS_MULTI", {})
    layer_schema = data.schema_map.get("LAYERS", {})
    fx_schema = data.schema_map.get("FX", {})
    character_schema = data.schema_map.get("CHARACTERS", {})

    if "TELP_PATTERNS" in wb.sheetnames:
        for r in iter_nonempty(load_sheet_dictionaries(wb["TELP_PATTERNS"])):
            pattern_id = _string_or_none(_row_value(r, telp_schema, "pattern_id", "パターンID"))
            if not pattern_id:
                continue
            data.telop_patterns[pattern_id] = TelopPattern(
                pattern_id=pattern_id,
                source=_string_or_none(_row_value(r, telp_schema, "source", "参照ソース")),
                overrides=parse_mapping(_row_value(r, telp_schema, "overrides", "上書きキー")),
                base_width=_safe_int(_row_value(r, telp_schema, "base_width", "基準幅")),
                base_height=_safe_int(_row_value(r, telp_schema, "base_height", "基準高さ")),
                fps=_safe_float(_row_value(r, telp_schema, "fps", "FPS")),
                description=_string_or_none(_row_value(r, telp_schema, "description", "説明")),
                notes=_string_or_none(_row_value(r, telp_schema, "notes", "備考")),
            )

    if "ASSETS_SINGLE" in wb.sheetnames:
        for r in iter_nonempty(load_sheet_dictionaries(wb["ASSETS_SINGLE"])):
            asset_id = _string_or_none(_row_value(r, asset_schema, "asset_id", "素材ID"))
            if not asset_id:
                continue
            parameters_raw = _row_value(r, asset_schema, "parameters", "パラメータ")
            data.assets[asset_id] = Asset(
                asset_id=asset_id,
                kind=_string_or_none(_row_value(r, asset_schema, "kind", "種別")),
                path=_string_or_none(_row_value(r, asset_schema, "path", "パス")),
                parameters=parse_mapping(parameters_raw) if parameters_raw is not None else {},
                default_layer=_safe_int(_row_value(r, asset_schema, "default_layer", "既定レイヤ")),
                default_x=_safe_float(_row_value(r, asset_schema, "default_x", "既定X")),
                default_y=_safe_float(_row_value(r, asset_schema, "default_y", "既定Y")),
                default_zoom=_safe_float(_row_value(r, asset_schema, "default_zoom", "既定ズーム")),
                notes=_string_or_none(_row_value(r, asset_schema, "notes", "備考")),
            )

    if "CHARACTERS" in wb.sheetnames:
        for r in iter_nonempty(load_sheet_dictionaries(wb["CHARACTERS"])):
            name = _string_or_none(_row_value(r, character_schema, "name", "キャラクター名"))
            part = _string_or_none(_row_value(r, character_schema, "part", "パーツ名"))
            base_path = _string_or_none(_row_value(r, character_schema, "base_path", "ベースパス"))
            if not (name and part and base_path):
                continue
            if name not in data.characters:
                data.characters[name] = Character(name=name)
            data.characters[name].parts[part] = base_path

    if "LAYERS" in wb.sheetnames:
        for r in iter_nonempty(load_sheet_dictionaries(wb["LAYERS"])):
            role = _string_or_none(_row_value(r, layer_schema, "role", "役割"))
            layer = _safe_int(_row_value(r, layer_schema, "layer", "レイヤ帯"))
            if role and layer is not None:
                data.layers[role] = LayerBand(role=role, layer=layer)

    if "PACKS_MULTI" in wb.sheetnames:
        for r in iter_nonempty(load_sheet_dictionaries(wb["PACKS_MULTI"])):
            pack_id = _string_or_none(_row_value(r, pack_schema, "pack_id", "パックID"))
            if not pack_id:
                continue
            data.packs[pack_id] = Pack(
                pack_id=pack_id,
                source=_string_or_none(_row_value(r, pack_schema, "source", "参照ソース")),
                overrides=parse_mapping(_row_value(r, pack_schema, "overrides", "上書きキー")),
                base_width=_safe_int(_row_value(r, pack_schema, "base_width", "基準幅")),
                base_height=_safe_int(_row_value(r, pack_schema, "base_height", "基準高さ")),
                fps=_safe_float(_row_value(r, pack_schema, "fps", "FPS")),
                notes=_string_or_none(_row_value(r, pack_schema, "notes", "備考")),
            )

    if "FX" in wb.sheetnames:
        for r in iter_nonempty(load_sheet_dictionaries(wb["FX"])):
            fx_id = _string_or_none(_row_value(r, fx_schema, "fx_id", "FX_ID"))
            if not fx_id:
                continue
            data.fx_presets[fx_id] = FxPreset(
                fx_id=fx_id,
                fx_type=_string_or_none(_row_value(r, fx_schema, "type", "種類")),
                source=_string_or_none(_row_value(r, fx_schema, "pack", "パック")),
                asset=_string_or_none(_row_value(r, fx_schema, "asset", "アセット")),
                parameters=parse_mapping(_row_value(r, fx_schema, "parameters", "パラメータ")),
            )

    # Resolve template paths for telops and assets
    for p in data.telop_patterns.values():
        _resolve_template_path(p, path)
    for a in data.assets.values():
        _resolve_template_path(a, path, "path", "parameters")
    for pack in data.packs.values():
        _resolve_template_path(pack, path)

    # Timeline
    if "TIMELINE" in wb.sheetnames:
        records = load_sheet_dictionaries(wb["TIMELINE"])
        timeline_schema = data.schema_map.get("TIMELINE", {})
        if records:
            keys = records[0].keys()
            column_key_map = _schema_column_key_map(timeline_schema)
            object_columns = _collect_schema_columns(timeline_schema, "object.")
            if not object_columns:
                object_columns = [k for k in keys if k and (k.startswith("オブジェクト") or k == "背景")]
            fx_columns = _collect_schema_columns(timeline_schema, "fx.")
            if not fx_columns:
                fx_columns = [k for k in keys if k and k.startswith("FX") and k != "FX_PARAM"]
            expr_config = {
                _single_column(timeline_schema, "expressions.primary", "表情(3つ内包)"): ["目", "口", "眉"],
                _single_column(timeline_schema, "expressions.face", "表情"): ["顔色"],
                _single_column(timeline_schema, "expressions.extra", "他1"): ["他1"],
            }
            expr_cols = {col: parts for col, parts in expr_config.items() if col}
            pack_columns = _schema_columns(timeline_schema, "packs", "パック")
            fx_param_columns = _schema_columns(timeline_schema, "fx.params", "FX_PARAM")
            approval_columns = _schema_columns(timeline_schema, "approval", "承認")
            memo_columns = _schema_columns(timeline_schema, "memo", "メモ")
            start_columns = _schema_columns(timeline_schema, "start", "開始")
            end_columns = _schema_columns(timeline_schema, "end", "終了")
            subtitle_columns = _schema_columns(timeline_schema, "subtitle", "字幕テキスト")
            telop_columns = _schema_columns(timeline_schema, "telop", "テロップ")
            character_columns = _schema_columns(timeline_schema, "character", "キャラクター")

            for i, r in enumerate(iter_nonempty(records), 1):
                expr = {
                    part: filename
                    for column, parts in expr_cols.items()
                    if column and (filename := _string_or_none(r.get(column)))
                    for part in parts
                }
                objs = []
                for column in object_columns:
                    identifier = _string_or_none(r.get(column))
                    if not identifier:
                        continue
                    canonical_key = column_key_map.get(column)
                    role_name = _resolve_role_name(column, canonical_key)
                    layer_band = _resolve_layer_band(data.layers, column, canonical_key, role_name)
                    objs.append(
                        TimelineObject(
                            role=role_name,
                            identifier=identifier,
                            layer=layer_band,
                            source_column=column,
                        )
                    )

                packs: List[str] = []
                for column in pack_columns:
                    packs.extend(
                        filter(
                            None,
                            (_normalize_identifier(value) for value in ensure_list(r.get(column))),
                        )
                    )

                fx_values: Dict[str, List[str]] = {}
                for column in fx_columns:
                    fx_values[column] = [
                        _normalize_identifier(value)
                        for value in ensure_list(r.get(column))
                        if _normalize_identifier(value)
                    ]

                total_fx = sum(len(values) for values in fx_values.values())
                fx_param_raw = _parse_fx_params(_first_nonempty(r, fx_param_columns))
                fxs: List[TimelineFx] = []
                for column_index, column in enumerate(fx_columns):
                    values = fx_values.get(column, [])
                    for fx_id in values:
                        params = _select_fx_parameters(fx_param_raw, fx_id, column, total_fx)
                        fxs.append(
                            TimelineFx(
                                fx_id=fx_id,
                                parameters=params,
                                source_column=column,
                                source_key=column_key_map.get(column),
                                column_index=column_index,
                            )
                        )

                notes: Dict[str, Any] = {}
                if fx_param_raw not in (None, {}):
                    notes["fx_params"] = fx_param_raw
                if approval := _string_or_none(_first_nonempty(r, approval_columns)):
                    notes["approval"] = approval
                if memo := _string_or_none(_first_nonempty(r, memo_columns)):
                    notes["memo"] = memo

                data.timeline.append(
                    TimelineRow(
                        index=i,
                        start=parse_timecode(_string_or_none(_first_nonempty(r, start_columns))),
                        end=parse_timecode(_string_or_none(_first_nonempty(r, end_columns))),
                        subtitle=_string_or_none(_first_nonempty(r, subtitle_columns)),
                        telop=_string_or_none(_first_nonempty(r, telop_columns)),
                        character=_string_or_none(_first_nonempty(r, character_columns)),
                        expressions=expr,
                        objects=objs,
                        fxs=fxs,
                        packs=[p for p in packs if p],
                        notes=notes,
                    )
                )
    return data

def _resolve_template_path(item: TelopPattern | Asset | Pack, wb_path: Path, path_field="source", param_field="overrides"):
    source_path_str = getattr(item, path_field)
    if source_path_str and Path(source_path_str).suffix == ".json":
        source_path = Path(source_path_str)
        if not source_path.is_absolute():
            source_path = wb_path.parent / source_path
        if source_path.exists():
            try:
                setattr(item, param_field, json.loads(source_path.read_text(encoding="utf-8-sig")))
            except (json.JSONDecodeError, IOError) as e:
                print(f"Warning: Failed to load template {source_path}: {e}")
        else:
            print(f"Warning: Template file not found: {source_path}")


def _safe_int(v):
    return int(v) if v is not None and v != "" else None


def _safe_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _string_or_none(v):
    if v is None or v == "":
        return None
    return str(v).strip()


def _normalize_identifier(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    text = str(value).strip()
    return text or None


def _parse_fx_params(raw: Any) -> Any:
    if raw in (None, ""):
        return None
    if isinstance(raw, (dict, list)):
        return raw
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return None
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            mapping = parse_mapping(text)
            return mapping if mapping else text
    return raw


def _normalize_param_value(value: Any) -> Dict[str, Any]:
    if value in (None, ""):
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        parsed = parse_mapping(value)
        return parsed if parsed else {"value": value}
    if isinstance(value, list):
        return {str(i): v for i, v in enumerate(value)}
    return {"value": value}


def _select_fx_parameters(raw: Any, fx_id: str, column_name: str, total_fx: int) -> Dict[str, Any]:
    if raw in (None, ""):
        return {}
    if isinstance(raw, dict):
        for key in (fx_id, column_name):
            if key in raw:
                return _normalize_param_value(raw[key])
        if total_fx == 1:
            return _normalize_param_value(raw)
        return {}
    if isinstance(raw, list):
        if total_fx == 1 and raw:
            return _normalize_param_value(raw[0])
        return {}
    if isinstance(raw, str):
        parsed = parse_mapping(raw)
        if parsed:
            if total_fx == 1:
                return parsed
            for key in (fx_id, column_name):
                if key in parsed:
                    return _normalize_param_value(parsed[key])
        return {"value": raw} if total_fx == 1 else {}
    return {"value": raw} if total_fx == 1 else {}
