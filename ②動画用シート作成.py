import sys
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from datetime import datetime
import traceback
from openpyxl import load_workbook # Excelファイルを読み込むために追加

# --- 準備 ---
script_path = Path(__file__).resolve()
project_root = script_path.parent
src_path = project_root / 'src'
sys.path.insert(0, str(src_path))
# srtファイルを解析する関数を直接インポート
from auto_movie_edit.srt import parse_srt, SrtParseError

# --- UIによるファイル選択処理 ---
def select_files():
    root = tk.Tk()
    root.withdraw()

    print("--- 台帳作成プロセスを開始します ---")
    
    srt_file = filedialog.askopenfilename(
        title="1. 読み込むSRTファイルを選択してください",
        filetypes=[("SRT Subtitles", "*.srt")]
    )
    if not srt_file:
        print("SRTファイルの選択がキャンセルされました。")
        return None, None
    
    print(f"✔️ SRTファイル: {Path(srt_file).name}")

    today_dir = project_root / datetime.now().strftime("%y%m%d")
    today_dir.mkdir(parents=True, exist_ok=True)

    workbook_file = filedialog.asksaveasfilename(
        title="2. 保存先のExcelファイル名を入力してください",
        filetypes=[("Excel Workbook", "*.xlsx")],
        defaultextension=".xlsx",
        initialfile="timeline.xlsx",
        initialdir=today_dir
    )
    if not workbook_file:
        print("保存先ファイルの選択がキャンセルされました。")
        return None, None

    workbook_path = today_dir / Path(workbook_file).name
    print(f"✔️ 保存先フォルダ: {today_dir}")

    return Path(srt_file), workbook_path

# --- 実行処理本体 (ここが新しいロジックです) ---
def main():
    srt_file, new_workbook_path = select_files()
    if not srt_file or not new_workbook_path:
        print("処理を中断しました。")
        return

    # テンプレートファイルのパスを定義
    template_path = project_root / "Template" / "template.xlsx"

    if not template_path.exists():
        print(f"エラー: テンプレートファイルが見つかりません。")
        print(f"次の場所に template.xlsx を配置してください: {template_path}")
        return

    try:
        # SRTファイルを解析
        print("SRTファイルを解析中...")
        entries = parse_srt(srt_file)

        # テンプレートExcelを読み込む
        print(f"テンプレート '{template_path.name}' を読み込み中...")
        workbook = load_workbook(template_path)
        timeline_sheet = workbook["TIMELINE"]

        # TIMELINEシートにSRTの内容を書き込む
        print("TIMELINEシートに書き込み中...")
        # 既存のデータをクリアせず、2行目から追記する
        start_row = timeline_sheet.max_row + 1
        for i, entry in enumerate(entries, start=start_row):
            timeline_sheet.cell(row=i, column=1, value=entry.start.to_string())
            timeline_sheet.cell(row=i, column=2, value=entry.end.to_string())
            timeline_sheet.cell(row=i, column=3, value=entry.text)

        # 新しいファイルとして保存
        workbook.save(new_workbook_path)
        
        print(f"\n--- 正常に処理が完了しました ---")
        print(f"テンプレートの内容を引き継いだ台帳が作成されました: {new_workbook_path}")

    except SrtParseError as e:
        print(f"\n--- SRTファイルの解析エラー ---")
        print(f"エラー内容: {e}")
    except KeyError:
        print(f"\n--- Excelシートのエラー ---")
        print(f"エラー: テンプレート '{template_path.name}' に 'TIMELINE' という名前のシートが見つかりません。")
    except Exception:
        print(f"\n--- 不明なエラーが発生しました ---")
        traceback.print_exc()

# --- 実行部分 ---
if __name__ == "__main__":
    main()
    print("\n--- Enterキーを押してウィンドウを閉じてください ---")
    input()