import sys
from pathlib import Path
import traceback
from openpyxl import load_workbook, Workbook

# --- 準備 ---
script_path = Path(__file__).resolve()
project_root = script_path.parent
src_path = project_root / 'src'
sys.path.insert(0, str(src_path))
from auto_movie_edit.workbook import DEFAULT_TEMPLATE, save_workbook, _write_headers

def update_workbook(workbook_path: Path):
    """
    Checks an existing workbook against the default template.
    Adds missing sheets and headers without deleting existing data.
    """
    if not workbook_path.exists():
        print(f"ファイルが見つかりません: {workbook_path}")
        print("新しいテンプレートファイルを作成します。")
        workbook = Workbook()
        workbook.remove(workbook.active) # Remove default sheet
    else:
        print(f"既存のファイルを開いています: {workbook_path.name}")
        workbook = load_workbook(workbook_path)

    sheets_to_check = {
        "TELP_PATTERNS": DEFAULT_TEMPLATE.telp_headers,
        "ASSETS_SINGLE": DEFAULT_TEMPLATE.asset_headers,
        "PACKS_MULTI": DEFAULT_TEMPLATE.pack_headers,
        "LAYERS": DEFAULT_TEMPLATE.layer_headers,
        "FX": DEFAULT_TEMPLATE.fx_headers,
        "TIMELINE": DEFAULT_TEMPLATE.timeline_headers,
        "SCHEMA_MAP": DEFAULT_TEMPLATE.schema_headers,
        "CHARACTERS": DEFAULT_TEMPLATE.characters_headers,
    }

    updated = False
    for sheet_name, headers in sheets_to_check.items():
        if sheet_name not in workbook.sheetnames:
            print(f"  > シート '{sheet_name}' を追加しています...")
            sheet = workbook.create_sheet(sheet_name)
            _write_headers(sheet, headers)
            updated = True
        else:
            # Check if headers are correct (optional, but good for completeness)
            sheet = workbook[sheet_name]
            if sheet.max_row == 0:
                 _write_headers(sheet, headers)
                 updated = True


    if updated:
        save_workbook(workbook, workbook_path)
        print("\n--- テンプレートの更新が完了しました ---")
    else:
        print("\n--- テンプレートは既に最新の状態です ---")


def main():
    """Main execution block."""
    print("--- Excelテンプレート更新ツール ---")
    template_path = project_root / "Template" / "template.xlsx"
    update_workbook(template_path)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        print("\n--- エラーが発生しました ---")
        traceback.print_exc()
    finally:
        print("\n--- Enterキーを押してウィンドウを閉じてください ---")
        input()